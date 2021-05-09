from openpyxl import load_workbook
import xlsxwriter

data_workbook = xlsxwriter.Workbook('Data_al_worked.xlsx')
worksheet = data_workbook.add_worksheet()

wb = load_workbook(r'C:\Users\Михаил\Desktop\VScode projects\php learning\HM\diploma\Data_alina.xlsx')
wb_ksk = load_workbook(r'C:\Users\Михаил\Desktop\VScode projects\php learning\HM\diploma\Data_alina_sec.xlsx')
print(wb.get_sheet_names())
sheet = wb.worksheets[0]
sheet2 = wb.worksheets[1]
sheet4 = wb.worksheets[3]
sheet5 = wb.worksheets[4]
sheetk = wb_ksk.worksheets[2]
for e in range(2, 57):
    worksheet.write(e, 0, sheet.cell(row=e, column=2).value)
    total_epst = 0
    kom_kos = 0
    org_kos = 0
    ais = 0
    scale_l = 0
    scale_a = 0
    scale_d = 0
    scale_k = 0
    scale_p = 0
    scale_m = 0
    scale_n = 0
    scale_v = 0
    scale_s = 0

    for i in range(8, 41):
        if i in [9, 13, 15, 16, 17, 18, 20, 22, 24, 25, 26, 27, 33, 34, 35, 37, 39]:
            if str(sheet.cell(row=e, column=i).value).startswith('Согл'):
                total_epst += 1
        if i in [10, 11, 12, 14, 19, 21, 23, 28, 29, 30, 31, 32, 36, 38, 40, 41]:
            if str(sheet.cell(row=e, column=i).value).startswith('Не согл'):
                total_epst += 1
    for i in range(8, 48):            
        if i in [9, 13, 17, 21, 25, 29, 33, 37, 41, 45]:
            if str(sheet2.cell(row=e, column=i).value).startswith('Согл'):
                kom_kos += 1
        if i in [11, 15, 19, 23, 27, 31, 35, 39, 43, 47]:
            if str(sheet2.cell(row=e, column=i).value).startswith('Не согл'):
                kom_kos += 1
        if i in [10, 14, 18, 22, 26, 30, 34, 38, 42, 46]:
            if str(sheet2.cell(row=e, column=i).value).startswith('Согл'):
                org_kos += 1
        if i in [12, 16, 20, 24, 28, 32, 36, 40, 44, 48]:
            if str(sheet2.cell(row=e, column=i).value).startswith('Не согл'):
                org_kos += 1

    for i in range(8, 33):
        if i in [14, 16, 20, 33]:
            if str(sheet4.cell(row=e, column=i).value).startswith('Не'):
                ais += 1
        else:
            if str(sheet4.cell(row=e, column=i).value).startswith('Да'):
                ais += 1

    for i in range(8, 108):
        if i in [19, 49, 89, 99]:
            if str(sheetk.cell(row=e, column=i).value).startswith('1'):
                scale_l += 1
        if i in [9, 29, 39, 59, 69, 79]:
            if str(sheetk.cell(row=e, column=i).value).startswith('3'):
                scale_l += 1
        if i in [10, 50, 60, 80, 100]:
            if str(sheetk.cell(row=e, column=i).value).startswith('1'):
                scale_a += 1
        if i in [20, 30, 40, 70, 90]:
            if str(sheetk.cell(row=e, column=i).value).startswith('3'):
                scale_a += 1
        if i in [11, 21, 31, 61, 91]:
            if str(sheetk.cell(row=e, column=i).value).startswith('1'):
                scale_d += 1
        if i in [41, 51, 71, 81, 101]:
            if str(sheetk.cell(row=e, column=i).value).startswith('3'):
                scale_d += 1
        if i in [22, 52, 72, 82, 92, 102]:
            if str(sheetk.cell(row=e, column=i).value).startswith('1'):
                scale_k += 1
        if i in [12, 32, 42, 62]:
            if str(sheetk.cell(row=e, column=i).value).startswith('3'):
                scale_k += 1
        if i in [13, 23, 43, 63, 93, 103]:
            if str(sheetk.cell(row=e, column=i).value).startswith('1'):
                scale_p += 1
        if i in [33, 53, 73, 83]:
            if str(sheetk.cell(row=e, column=i).value).startswith('3'):
                scale_p += 1
        if i in [24, 34, 54, 64, 94]:
            if str(sheetk.cell(row=e, column=i).value).startswith('1'):
                scale_m += 1
        if i in [14, 44, 74, 84, 104]:
            if str(sheetk.cell(row=e, column=i).value).startswith('3'):
                scale_m += 1
        if i in [25, 35, 45, 85, 105]:
            if str(sheetk.cell(row=e, column=i).value).startswith('1'):
                scale_n += 1
        if i in [15, 55, 65, 75, 95]:
            if str(sheetk.cell(row=e, column=i).value).startswith('3'):
                scale_n += 1
        if i in [16, 26, 36, 56, 76, 96]:
            if str(sheetk.cell(row=e, column=i).value).startswith('2'):
                scale_v += 1
        if i in [46, 66, 86, 106]:
            if str(sheetk.cell(row=e, column=i).value).startswith('3'):
                scale_v += 1
        if i in [28, 38, 58, 68, 88, 108]:
            if str(sheetk.cell(row=e, column=i).value).startswith('1'):
                scale_s += 1
        if i in [18, 48, 78, 98]:
            if str(sheetk.cell(row=e, column=i).value).startswith('3'):
                scale_s += 1
        




        
        worksheet.write(e, 2, total_epst)
        worksheet.write(e, 3, kom_kos)
        worksheet.write(e, 4, org_kos)
        worksheet.write(e, 5, ais)
        worksheet.write(e, 6, scale_l)
        worksheet.write(e, 7, scale_a)
        worksheet.write(e, 8, scale_d)
        worksheet.write(e, 9, scale_k)
        worksheet.write(e, 10, scale_p)
        worksheet.write(e, 11, scale_m)
        worksheet.write(e, 12, scale_n)
        worksheet.write(e, 13, scale_v)
        worksheet.write(e, 14, scale_s)
worksheet.write(1,2, 'Эпштейн')
worksheet.write(1,3, 'КОС Коммуникативность')
worksheet.write(1,4, 'КОС Организационные способности')
worksheet.write(1,5, 'Айзенк (э-и)')
worksheet.write(1,6, 'Шкала Л')
worksheet.write(1,7, 'Шкала А')
worksheet.write(1,8, 'Шкала Д')
worksheet.write(1,9, 'Шкала К')
worksheet.write(1,10, 'Шкала П')
worksheet.write(1,11, 'Шкала М')
worksheet.write(1,12, 'Шкала Н')
worksheet.write(1,13, 'Шкала В')
worksheet.write(1,14, 'Шкала С')

data_workbook.close()
