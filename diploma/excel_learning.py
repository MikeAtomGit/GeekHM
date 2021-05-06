from openpyxl import load_workbook
import xlsxwriter

data_workbook = xlsxwriter.Workbook('Data.xlsx')
worksheet = data_workbook.add_worksheet()

wb = load_workbook(r'C:\Users\Михаил\Desktop\VScode projects\php learning\HM\diploma\Diplom__1.xlsx')
print(wb.get_sheet_names())
sheet = wb.worksheets[0]
for e in range(2, 57):
    worksheet.write(e, 0, sheet.cell(row=e, column=2).value)
    score1 = 0
    score2_1 = 0
    score2_2 = 0
    score2_3 = 0
    score3_1 = 0
    score3_2 = 0
    score3_3 = 0
    score3_4 = 0
    score3_5 = 0
    score3_6 = 0
    for i in range(5, 96):
        if i in [5, 7, 8, 9, 13, 14, 15, 18, 19, 22, 23, 24, 26, 28, 29]:
            score1 = int(sheet.cell(row=e, column=i).value) + score1
        elif i in [6, 10, 11, 12, 16, 17, 20, 21, 25, 27]:
            score1 += (8 - int(sheet.cell(row=e, column=i).value))
        worksheet.write(e,1, score1)
        
        if str(sheet.cell(row=e, column=i).value).startswith('а)'):
            if i in [32, 33, 37, 39, 44, 45, 49, 50, 53, 54]:
                score2_1 += 1
            if i in [36, 38, 43, 48, 56, 57, 58]:
                score2_2 += 1
            if i in [34, 35, 40, 41, 42, 46,  47, 51, 52]:
                score2_3 += 1
                
        if str(sheet.cell(row=e, column=i).value).startswith('б)'):
            if i in [35, 38, 41, 42, 43, 47, 52, 58]:
                score2_1 += 1
            if i in [32, 34, 36, 40, 46, 48, 51, 55, 57]:
                score2_2 += 1
            if i in [33, 37, 39, 44, 45, 49, 50, 53, 54, 56]:
                score2_3 += 1

        if str(sheet.cell(row=e, column=i).value).startswith('в)'):
            if i in [33, 34, 45, 46, 50, 54, 56, 57]:
                score2_1 += 1
            if i in [32, 37, 38, 39, 40, 44, 49, 51, 53]:
                score2_2 += 1
            if i in [35, 36, 41, 42, 43, 47, 48, 52, 55, 58]:
                score2_3 += 1
                
        if str(sheet.cell(row=e, column=i).value).startswith('г)'):
            if i in [32, 36, 37, 38, 39, 43, 44, 48, 49, 51, 55]:
                score2_1 += 1
            if i in [35, 41, 42, 47, 52, 53, 58]:
                score2_2 += 1
            if i in [33, 34, 40, 45, 46, 50, 54, 56, 57]:
                score2_3 += 1
                
        if str(sheet.cell(row=e, column=i).value).startswith('д)'):
            if i in [34, 35, 40, 41, 42, 46, 47, 51, 57, 58]:
                score2_1 += 1
            if i in [33, 37, 45, 50, 52, 54, 56]:
                score2_2 += 1
            if i in [32, 36, 38, 39, 43, 44, 48, 49, 53, 55]:
                score2_3 += 1

        if i in [59, 65, 77, 83]:
            if str(sheet.cell(row=e, column=i).value).startswith('да'):
                score3_1 += 1
        if i in [71, 89]:
            if str(sheet.cell(row=e, column=i).value).startswith('Н'):
                score3_1 += 1
        if i in [66, 78, 90]:
            if str(sheet.cell(row=e, column=i).value).startswith('да'):
                score3_2 += 1
        if i in [60, 72, 84]:
             if str(sheet.cell(row=e, column=i).value).startswith('Н'):
                score3_2 += 1
        if i in [67, 73, 79, 86]:
            if str(sheet.cell(row=e, column=i).value).startswith('да'):
                score3_3 += 1
        if i in [61, 91]:        
            if str(sheet.cell(row=e, column=i).value).startswith('Н'):
                score3_3 += 1
        if i in [62]:
            if str(sheet.cell(row=e, column=i).value).startswith('да'):
                score3_4 += 1
        if i in [68, 74, 80, 86, 92]:
            if str(sheet.cell(row=e, column=i).value).startswith('Н'):
                score3_4 += 1
        if i in [63]:
            if str(sheet.cell(row=e, column=i).value).startswith('да'):
                score3_5 += 1
        if i in [69, 75, 81, 87, 93]:
            if str(sheet.cell(row=e, column=i).value).startswith('Н'):
                score3_5 += 1
        if i in [94]:
            if str(sheet.cell(row=e, column=i).value).startswith('Н'):
                score3_6 += 1
        if i in [64, 70, 76, 82, 88]:
            if str(sheet.cell(row=e, column=i).value).startswith('да'):
                score3_6 += 1

        
        worksheet.write(e, 2, score2_1)
        worksheet.write(e, 3, score2_2)
        worksheet.write(e, 4, score2_3)
        worksheet.write(e, 5, score3_1)
        worksheet.write(e, 6, score3_2)
        worksheet.write(e, 7, score3_3)
        worksheet.write(e, 8, score3_4)
        worksheet.write(e, 9, score3_5)
        worksheet.write(e, 10, score3_6)
worksheet.write(1,2, 'зависимые')
worksheet.write(1,3, 'компетентные')
worksheet.write(1,4, 'агрессивные')
worksheet.write(1,5, 'Рациональный канал эмпатии')
worksheet.write(1,6, 'Эмоциональный канал эмпатии')
worksheet.write(1,7, 'Интуитивный канал эмпатии')
worksheet.write(1,8, 'Установки, способствующие эмпатии')
worksheet.write(1,9, 'Проникающая способность в эмпатии')
worksheet.write(1,10, 'Идентификация в эмпатии')

data_workbook.close()
