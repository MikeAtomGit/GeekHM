from openpyxl import load_workbook
import xlsxwriter
import checka as ffs
import re

data_workbook = xlsxwriter.Workbook('xldip_worked.xlsx')
worksheet = data_workbook.add_worksheet()

wb = load_workbook(r'C:\Users\Михаил\Desktop\VScode projects\php learning\HM\diploma\xldip.xlsx')
print(wb.get_sheet_names())
sheet = wb.worksheets[0]

worksheet.write(1, 2, 'Катастрофизация')
worksheet.write(1, 3, 'Долженствование в отношении себя')
worksheet.write(1, 4, 'Долженствование в отношении других')
worksheet.write(1, 5, 'Самооценка и рациональность')
worksheet.write(1, 6, 'Фрустрированность')
worksheet.write(1, 8, 'Гипертимный')
worksheet.write(1, 9, 'Циклоидный')
worksheet.write(1, 10, 'Лабильный')
worksheet.write(1, 11, 'Астено-невротический')
worksheet.write(1, 12, 'Сенситивный')
worksheet.write(1, 13, 'Психастенический')
worksheet.write(1, 14, 'Шизоидный')
worksheet.write(1, 15, 'Эпилептоидный')
worksheet.write(1, 16, 'Истероидный')
worksheet.write(1, 17, 'Неустойчивый')
worksheet.write(1, 18, 'Конформный')
worksheet.write(1, 19, 'Диссимуляция')
worksheet.write(1, 20, 'Откровенность')
worksheet.write(1, 21, 'Потенциальная психопатия')
worksheet.write(1, 22, 'Эмансипация')
worksheet.write(1, 23, 'Делинквентность (М - П) (НЕ УЧИТЫВАТЬ)')
worksheet.write(1, 24, 'Мужественность')
worksheet.write(1, 25, 'Феминность')

def cmp(a, b):
    return [c for c in a if c.isalpha()] == [c for c in b if c.isalpha()]
def split_uppercase(value):
    return re.split('[ ]{2}', re.sub(r'([А-Я])', r' \1', value))

for e in range(2,3):
    worksheet.write(e, 0, sheet.cell(row=e, column=2).value)
    kat = 0
    dol_s = 0
    dol_o = 0
    sam = 0
    fru = 0
    s = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
    
    
    for i in range(54, 200):
        
        if str(sheet.cell(row=e, column=i).value).startswith('Полностью согласен'):
            if i in [9, 14, 19, 24, 34, 39, 44]:
                kat += 1
            if i in [4, 29, 49]:
                kat += 3
            if i in [5, 10, 15, 30, 35, 40, 50]:
                dol_s += 1
            if i in [20, 25, 45]:
                dol_s += 3
            if i in [6, 11, 21, 26, 36, 46, 51]:
                dol_o += 1
            if i in [12, 27, 37]:
                dol_o += 3
            if i in [8, 13, 18, 33, 38, 43, 48, 53]:
                sam += 1
            if i in [20, 25]:
                sam += 3
            if i in [12, 17, 22, 27, 32, 42, 47]:
                fru += 1
            if i in [4, 34, 49]:
                fru += 3
        if str(sheet.cell(row=e, column=i).value).startswith('Не'):
            if i in [9, 14, 19, 24, 34, 39, 44]:
                kat += 2
            if i in [4, 29, 49]:
                kat += 2
            if i in [5, 10, 15, 30, 35, 40, 50]:
                dol_s += 2
            if i in [20, 25, 45]:
                dol_s += 2
            if i in [6, 11, 21, 26, 36, 46, 51]:
                dol_o += 2
            if i in [12, 27, 37]:
                dol_o += 2
            if i in [8, 13, 18, 33, 38, 43, 48, 53]:
                sam += 2
            if i in [20, 25]:
                sam += 2
            if i in [12, 17, 22, 27, 32, 42, 47]:
                fru += 2
            if i in [4, 34, 49]:
                fru += 2
        if str(sheet.cell(row=e, column=i).value).startswith('Полностью не'):
            if i in [9, 14, 19, 24, 34, 39, 44]:
                kat += 3
            if i in [4, 29, 49]:
                kat += 1
            if i in [5, 10, 15, 30, 35, 40, 50]:
                dol_s += 3
            if i in [20, 25, 45]:
                dol_s += 1
            if i in [6, 11, 21, 26, 36, 46, 51]:
                dol_o += 3
            if i in [12, 27, 37]:
                dol_o += 1
            if i in [8, 13, 18, 33, 38, 43, 48, 53]:
                sam += 3
            if i in [20, 25]:
                sam += 1
            if i in [12, 17, 22, 27, 32, 42, 47]:
                fru += 3
            if i in [4, 34, 49]:
                fru += 1
        
        if i in range(54, 79):
            x = split_uppercase(str(sheet.cell(row=e, column=i).value))
            s = ffs.count_fucking_result(x, s)
            print(s)
            
            
        if i in range(79, 103):
            x = split_uppercase(str(sheet.cell(row=e, column=i).value))
            s = ffs.count_fucking_result1(x, s)
            print(s)
            
        print(s)
        
        worksheet.write(e, 2, kat)
        worksheet.write(e, 3, dol_s)
        worksheet.write(e, 4, dol_o)
        worksheet.write(e, 5, sam)
        worksheet.write(e, 6, fru)
        
        worksheet.write(e, 8, s[0])
        worksheet.write(e, 9, s[1])
        worksheet.write(e, 10, s[2])
        worksheet.write(e, 11, s[3])
        worksheet.write(e, 12, s[4])
        worksheet.write(e, 13, s[5])
        worksheet.write(e, 14, s[6])
        worksheet.write(e, 15, s[7])
        worksheet.write(e, 16, s[8])
        worksheet.write(e, 17, s[9])
        worksheet.write(e, 18, s[10])
        worksheet.write(e, 19, s[11])
        worksheet.write(e, 20, s[12])
        worksheet.write(e, 21, s[13])
        worksheet.write(e, 22, s[14])
        worksheet.write(e, 23, s[15])
        worksheet.write(e, 24, s[16])
        worksheet.write(e, 25, s[17])
        
        
        
        

data_workbook.close()
